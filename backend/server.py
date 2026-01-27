from fastapi import FastAPI, HTTPException, Depends, Request, Response, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone, timedelta
from dotenv import load_dotenv
import os
import uuid
import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from auth_utils import verify_password, get_password_hash, create_access_token, decode_access_token

load_dotenv()

app = FastAPI()

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB connection
MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
client = AsyncIOMotorClient(MONGO_URL)
db = client.test_database

# Pydantic Models
class User(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: str = "user"  # user, approver, admin
    created_at: datetime

class SessionDataResponse(BaseModel):
    id: str
    email: str
    name: str
    picture: Optional[str] = None
    session_token: str

class CostCenter(BaseModel):
    center_id: str
    name: str
    code: str
    active: bool = True
    created_at: datetime

class ExpenseCategory(BaseModel):
    category_id: str
    name: str
    active: bool = True
    order: int
    created_at: datetime

class Trip(BaseModel):
    trip_id: str
    name: str
    creator_id: str
    cost_center_id: str
    status: str = "pending"  # pending, approved, rejected
    participants: List[str]
    created_at: datetime
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    rejection_reason: Optional[str] = None

class Expense(BaseModel):
    expense_id: str
    trip_id: str
    user_id: str
    amount: float
    date: datetime
    establishment: str
    category_id: str
    receipt_image: str  # base64
    notes: Optional[str] = None
    created_at: datetime
    modified_at: datetime
    modified_by: str

class AuditLog(BaseModel):
    log_id: str
    entity_type: str  # trip, expense
    entity_id: str
    action: str  # created, updated, deleted, approved, rejected
    user_id: str
    changes: Dict[str, Any]
    timestamp: datetime

# Request/Response Models
class LoginRequest(BaseModel):
    email: str
    password: str

class RegisterRequest(BaseModel):
    email: str
    password: str
    name: str

class CreateUserRequest(BaseModel):
    email: str
    name: str
    role: str = "user"

class UpdateUserRoleRequest(BaseModel):
    role: str

class CreateCostCenterRequest(BaseModel):
    name: str
    code: str

class UpdateCostCenterRequest(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    active: Optional[bool] = None

class CreateExpenseCategoryRequest(BaseModel):
    name: str
    order: Optional[int] = None

class UpdateExpenseCategoryRequest(BaseModel):
    name: Optional[str] = None
    active: Optional[bool] = None
    order: Optional[int] = None

class CreateTripRequest(BaseModel):
    name: str
    cost_center_id: str
    participants: List[str]

class UpdateTripRequest(BaseModel):
    name: Optional[str] = None
    cost_center_id: Optional[str] = None
    participants: Optional[List[str]] = None

class ApproveRejectRequest(BaseModel):
    reason: Optional[str] = None

class CreateExpenseRequest(BaseModel):
    trip_id: str
    amount: float
    date: str  # ISO format
    establishment: str
    category_id: str
    receipt_image: str  # base64
    notes: Optional[str] = None

class UpdateExpenseRequest(BaseModel):
    amount: Optional[float] = None
    date: Optional[str] = None
    establishment: Optional[str] = None
    category_id: Optional[str] = None
    receipt_image: Optional[str] = None
    notes: Optional[str] = None

# Auth Helper
async def get_current_user(
    request: Request,
    authorization: Optional[str] = Header(None)
) -> User:
    """Get current user from JWT token"""
    token = None
    
    # Try to get token from Authorization header
    if authorization and authorization.startswith("Bearer "):
        token = authorization.replace("Bearer ", "")
    
    # Try to get token from cookie
    if not token:
        token = request.cookies.get("access_token")
    
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Decode JWT token
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    user_id = payload.get("user_id")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token payload")
    
    # Get user from database
    user_doc = await db.users.find_one(
        {"user_id": user_id},
        {"_id": 0, "password_hash": 0}  # Exclude password hash
    )
    
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    
    return User(**user_doc)

async def require_role(user: User, allowed_roles: List[str]):
    if user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

async def create_audit_log(
    entity_type: str,
    entity_id: str,
    action: str,
    user_id: str,
    changes: Dict[str, Any]
):
    log = {
        "log_id": f"log_{uuid.uuid4().hex[:12]}",
        "entity_type": entity_type,
        "entity_id": entity_id,
        "action": action,
        "user_id": user_id,
        "changes": changes,
        "timestamp": datetime.now(timezone.utc)
    }
    await db.audit_logs.insert_one(log)

# Auth Endpoints
@app.post("/api/auth/register")
async def register(data: RegisterRequest):
    """Register a new user (public endpoint)"""
    try:
        # Check if user already exists
        existing = await db.users.find_one({"email": data.email}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail="Email already registered")
        
        # Create user
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        password_hash = get_password_hash(data.password)
        
        user = {
            "user_id": user_id,
            "email": data.email,
            "name": data.name,
            "password_hash": password_hash,
            "picture": None,
            "role": "user",
            "created_at": datetime.now(timezone.utc)
        }
        
        await db.users.insert_one(user)
        
        # Create JWT token
        access_token = create_access_token({"user_id": user_id})
        
        # Remove password_hash from response
        user.pop("password_hash")
        user.pop("_id", None)
        
        return {
            "user": user,
            "access_token": access_token,
            "token_type": "bearer"
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Registration error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/auth/login")
async def login(data: LoginRequest, response: Response):
    """Login with email and password"""
    try:
        # Find user
        user_doc = await db.users.find_one({"email": data.email}, {"_id": 0})
        
        if not user_doc:
            raise HTTPException(status_code=401, detail="Invalid email or password")
        
        # Verify password
        password_hash = user_doc.get("password_hash")
        if not password_hash or not verify_password(data.password, password_hash):
            raise HTTPException(status_code=401, detail="Invalid email or password")
        
        # Create JWT token
        access_token = create_access_token({"user_id": user_doc["user_id"]})
        
        # Set cookie
        response.set_cookie(
            key="access_token",
            value=access_token,
            httponly=True,
            secure=True,
            samesite="none",
            max_age=7 * 24 * 60 * 60,  # 7 days
            path="/"
        )
        
        # Remove password_hash from response
        user_doc.pop("password_hash", None)
        
        return {
            "user": user_doc,
            "access_token": access_token,
            "token_type": "bearer"
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Login error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

@app.post("/api/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"message": "Logged out"}

# Admin Endpoints - User Management
@app.post("/api/admin/users")
async def create_user(
    data: CreateUserRequest,
    current_user: User = Depends(get_current_user)
):
    await require_role(current_user, ["admin"])
    
    # Check if user exists
    existing = await db.users.find_one({"email": data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="User already exists")
    
    # Create temporary password (user should change it on first login)
    temp_password = "Password123!"  # En producción, generar contraseña aleatoria y enviar por email
    password_hash = get_password_hash(temp_password)
    
    user = {
        "user_id": f"user_{uuid.uuid4().hex[:12]}",
        "email": data.email,
        "name": data.name,
        "password_hash": password_hash,
        "picture": None,
        "role": data.role,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.users.insert_one(user)
    return {
        "message": "User created",
        "user_id": user["user_id"],
        "temporary_password": temp_password
    }

@app.get("/api/admin/users")
async def list_users(current_user: User = Depends(get_current_user)):
    await require_role(current_user, ["admin"])
    
    users = await db.users.find({}, {"_id": 0}).to_list(1000)
    return {"users": users}

@app.put("/api/admin/users/{user_id}")
async def update_user_role(
    user_id: str,
    data: UpdateUserRoleRequest,
    current_user: User = Depends(get_current_user)
):
    await require_role(current_user, ["admin"])
    
    result = await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": data.role}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User role updated"}

# Admin Endpoints - Cost Centers
@app.post("/api/admin/cost-centers")
async def create_cost_center(
    data: CreateCostCenterRequest,
    current_user: User = Depends(get_current_user)
):
    await require_role(current_user, ["admin"])
    
    center = {
        "center_id": f"center_{uuid.uuid4().hex[:12]}",
        "name": data.name,
        "code": data.code,
        "active": True,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.cost_centers.insert_one(center)
    return {"message": "Cost center created", "center_id": center["center_id"]}

@app.get("/api/admin/cost-centers")
async def list_cost_centers(current_user: User = Depends(get_current_user)):
    centers = await db.cost_centers.find({}, {"_id": 0}).to_list(1000)
    return {"cost_centers": centers}

@app.put("/api/admin/cost-centers/{center_id}")
async def update_cost_center(
    center_id: str,
    data: UpdateCostCenterRequest,
    current_user: User = Depends(get_current_user)
):
    await require_role(current_user, ["admin"])
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = await db.cost_centers.update_one(
        {"center_id": center_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cost center not found")
    
    return {"message": "Cost center updated"}

# Admin Endpoints - Expense Categories
@app.post("/api/admin/expense-categories")
async def create_expense_category(
    data: CreateExpenseCategoryRequest,
    current_user: User = Depends(get_current_user)
):
    await require_role(current_user, ["admin"])
    
    # Get max order if not provided
    order = data.order
    if order is None:
        max_cat = await db.expense_categories.find_one(
            {},
            {"_id": 0, "order": 1},
            sort=[("order", -1)]
        )
        order = (max_cat["order"] + 1) if max_cat else 1
    
    category = {
        "category_id": f"category_{uuid.uuid4().hex[:12]}",
        "name": data.name,
        "active": True,
        "order": order,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.expense_categories.insert_one(category)
    return {"message": "Category created", "category_id": category["category_id"]}

@app.get("/api/admin/expense-categories")
async def list_expense_categories(current_user: User = Depends(get_current_user)):
    categories = await db.expense_categories.find(
        {},
        {"_id": 0}
    ).sort("order", 1).to_list(1000)
    return {"categories": categories}

@app.put("/api/admin/expense-categories/{category_id}")
async def update_expense_category(
    category_id: str,
    data: UpdateExpenseCategoryRequest,
    current_user: User = Depends(get_current_user)
):
    await require_role(current_user, ["admin"])
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = await db.expense_categories.update_one(
        {"category_id": category_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    return {"message": "Category updated"}

# Trip Endpoints
@app.post("/api/trips")
async def create_trip(
    data: CreateTripRequest,
    current_user: User = Depends(get_current_user)
):
    # Verify cost center exists
    center = await db.cost_centers.find_one({"center_id": data.cost_center_id}, {"_id": 0})
    if not center:
        raise HTTPException(status_code=404, detail="Cost center not found")
    
    # Ensure creator is in participants
    participants = list(set([current_user.user_id] + data.participants))
    
    trip = {
        "trip_id": f"trip_{uuid.uuid4().hex[:12]}",
        "name": data.name,
        "creator_id": current_user.user_id,
        "cost_center_id": data.cost_center_id,
        "status": "pending",
        "participants": participants,
        "created_at": datetime.now(timezone.utc),
        "approved_by": None,
        "approved_at": None,
        "rejection_reason": None
    }
    
    await db.trips.insert_one(trip)
    
    # Audit log
    await create_audit_log(
        "trip",
        trip["trip_id"],
        "created",
        current_user.user_id,
        {"trip": trip}
    )
    
    return {"message": "Trip created", "trip_id": trip["trip_id"], "trip": trip}

@app.get("/api/trips")
async def list_my_trips(current_user: User = Depends(get_current_user)):
    trips = await db.trips.find(
        {"participants": current_user.user_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    return {"trips": trips}

@app.get("/api/trips/pending")
async def list_pending_trips(current_user: User = Depends(get_current_user)):
    await require_role(current_user, ["approver", "admin"])
    
    trips = await db.trips.find(
        {"status": "pending"},
        {"_id": 0}
    ).sort("created_at", 1).to_list(1000)
    
    return {"trips": trips}

@app.get("/api/trips/{trip_id}")
async def get_trip(
    trip_id: str,
    current_user: User = Depends(get_current_user)
):
    trip = await db.trips.find_one({"trip_id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Check access
    if current_user.user_id not in trip["participants"] and current_user.role not in ["approver", "admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return trip

@app.put("/api/trips/{trip_id}")
async def update_trip(
    trip_id: str,
    data: UpdateTripRequest,
    current_user: User = Depends(get_current_user)
):
    trip = await db.trips.find_one({"trip_id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Only creator can update and only if pending
    if trip["creator_id"] != current_user.user_id:
        raise HTTPException(status_code=403, detail="Only creator can update")
    
    if trip["status"] != "pending":
        raise HTTPException(status_code=400, detail="Cannot update approved/rejected trip")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    # Ensure creator stays in participants
    if "participants" in update_data:
        update_data["participants"] = list(set([current_user.user_id] + update_data["participants"]))
    
    result = await db.trips.update_one(
        {"trip_id": trip_id},
        {"$set": update_data}
    )
    
    # Audit log
    await create_audit_log(
        "trip",
        trip_id,
        "updated",
        current_user.user_id,
        {"changes": update_data}
    )
    
    return {"message": "Trip updated"}

@app.post("/api/trips/{trip_id}/approve")
async def approve_trip(
    trip_id: str,
    current_user: User = Depends(get_current_user)
):
    await require_role(current_user, ["approver", "admin"])
    
    trip = await db.trips.find_one({"trip_id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip["status"] != "pending":
        raise HTTPException(status_code=400, detail="Trip is not pending")
    
    result = await db.trips.update_one(
        {"trip_id": trip_id},
        {
            "$set": {
                "status": "approved",
                "approved_by": current_user.user_id,
                "approved_at": datetime.now(timezone.utc)
            }
        }
    )
    
    # Audit log
    await create_audit_log(
        "trip",
        trip_id,
        "approved",
        current_user.user_id,
        {"approved_by": current_user.user_id}
    )
    
    return {"message": "Trip approved"}

@app.post("/api/trips/{trip_id}/reject")
async def reject_trip(
    trip_id: str,
    data: ApproveRejectRequest,
    current_user: User = Depends(get_current_user)
):
    await require_role(current_user, ["approver", "admin"])
    
    trip = await db.trips.find_one({"trip_id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip["status"] != "pending":
        raise HTTPException(status_code=400, detail="Trip is not pending")
    
    result = await db.trips.update_one(
        {"trip_id": trip_id},
        {
            "$set": {
                "status": "rejected",
                "approved_by": current_user.user_id,
                "approved_at": datetime.now(timezone.utc),
                "rejection_reason": data.reason
            }
        }
    )
    
    # Audit log
    await create_audit_log(
        "trip",
        trip_id,
        "rejected",
        current_user.user_id,
        {"rejected_by": current_user.user_id, "reason": data.reason}
    )
    
    return {"message": "Trip rejected"}

# Expense Endpoints
@app.post("/api/expenses")
async def create_expense(
    data: CreateExpenseRequest,
    current_user: User = Depends(get_current_user)
):
    # Verify trip exists and is approved
    trip = await db.trips.find_one({"trip_id": data.trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip["status"] != "approved":
        raise HTTPException(status_code=400, detail="Trip is not approved")
    
    # Verify user is participant
    if current_user.user_id not in trip["participants"]:
        raise HTTPException(status_code=403, detail="Not a trip participant")
    
    # Verify category exists
    category = await db.expense_categories.find_one({"category_id": data.category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    expense = {
        "expense_id": f"expense_{uuid.uuid4().hex[:12]}",
        "trip_id": data.trip_id,
        "user_id": current_user.user_id,
        "amount": data.amount,
        "date": datetime.fromisoformat(data.date.replace('Z', '+00:00')),
        "establishment": data.establishment,
        "category_id": data.category_id,
        "receipt_image": data.receipt_image,
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc),
        "modified_at": datetime.now(timezone.utc),
        "modified_by": current_user.user_id
    }
    
    await db.expenses.insert_one(expense)
    
    # Audit log
    await create_audit_log(
        "expense",
        expense["expense_id"],
        "created",
        current_user.user_id,
        {"expense": expense}
    )
    
    return {"message": "Expense created", "expense_id": expense["expense_id"]}

@app.get("/api/expenses/trip/{trip_id}")
async def list_trip_expenses(
    trip_id: str,
    current_user: User = Depends(get_current_user)
):
    # Verify trip access
    trip = await db.trips.find_one({"trip_id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if current_user.user_id not in trip["participants"] and current_user.role not in ["approver", "admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    expenses = await db.expenses.find(
        {"trip_id": trip_id},
        {"_id": 0}
    ).sort("date", -1).to_list(1000)
    
    return {"expenses": expenses}

@app.get("/api/expenses/{expense_id}")
async def get_expense(
    expense_id: str,
    current_user: User = Depends(get_current_user)
):
    expense = await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    # Check access via trip
    trip = await db.trips.find_one({"trip_id": expense["trip_id"]}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if current_user.user_id not in trip["participants"] and current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return expense

@app.put("/api/expenses/{expense_id}")
async def update_expense(
    expense_id: str,
    data: UpdateExpenseRequest,
    current_user: User = Depends(get_current_user)
):
    expense = await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    # Check permissions: owner or admin
    if expense["user_id"] != current_user.user_id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    # Parse date if provided
    if "date" in update_data:
        update_data["date"] = datetime.fromisoformat(update_data["date"].replace('Z', '+00:00'))
    
    update_data["modified_at"] = datetime.now(timezone.utc)
    update_data["modified_by"] = current_user.user_id
    
    result = await db.expenses.update_one(
        {"expense_id": expense_id},
        {"$set": update_data}
    )
    
    # Audit log
    await create_audit_log(
        "expense",
        expense_id,
        "updated",
        current_user.user_id,
        {"changes": update_data}
    )
    
    return {"message": "Expense updated"}

@app.delete("/api/expenses/{expense_id}")
async def delete_expense(
    expense_id: str,
    current_user: User = Depends(get_current_user)
):
    expense = await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    # Check permissions: owner or admin
    if expense["user_id"] != current_user.user_id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.expenses.delete_one({"expense_id": expense_id})
    
    # Audit log
    await create_audit_log(
        "expense",
        expense_id,
        "deleted",
        current_user.user_id,
        {"expense": expense}
    )
    
    return {"message": "Expense deleted"}

# Audit Endpoints
@app.get("/api/audit/{entity_type}/{entity_id}")
async def get_audit_logs(
    entity_type: str,
    entity_id: str,
    current_user: User = Depends(get_current_user)
):
    logs = await db.audit_logs.find(
        {"entity_type": entity_type, "entity_id": entity_id},
        {"_id": 0}
    ).sort("timestamp", -1).to_list(1000)
    
    return {"logs": logs}

# Excel Export
@app.get("/api/admin/export/excel")
async def export_to_excel(
    trip_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    await require_role(current_user, ["admin"])
    
    # Build query
    query = {}
    if trip_id:
        query["trip_id"] = trip_id
    
    # Get all expenses
    expenses = await db.expenses.find(query, {"_id": 0}).to_list(10000)
    
    if not expenses:
        raise HTTPException(status_code=404, detail="No expenses found")
    
    # Get related data
    user_ids = list(set([e["user_id"] for e in expenses]))
    trip_ids = list(set([e["trip_id"] for e in expenses]))
    category_ids = list(set([e["category_id"] for e in expenses]))
    
    users = await db.users.find({"user_id": {"$in": user_ids}}, {"_id": 0}).to_list(1000)
    trips = await db.trips.find({"trip_id": {"$in": trip_ids}}, {"_id": 0}).to_list(1000)
    categories = await db.expense_categories.find({"category_id": {"$in": category_ids}}, {"_id": 0}).to_list(1000)
    
    # Create lookups
    user_map = {u["user_id"]: u["name"] for u in users}
    trip_map = {t["trip_id"]: t for t in trips}
    category_map = {c["category_id"]: c["name"] for c in categories}
    
    # Get cost centers
    center_ids = list(set([t["cost_center_id"] for t in trips]))
    centers = await db.cost_centers.find({"center_id": {"$in": center_ids}}, {"_id": 0}).to_list(1000)
    center_map = {c["center_id"]: c["name"] for c in centers}
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos de Viaje"
    
    # Headers
    headers = ["Usuario", "Centro de Coste", "Fecha", "Gasto", "Motivo", "Establecimiento", "Viaje"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, expense in enumerate(expenses, 2):
        trip = trip_map.get(expense["trip_id"], {})
        cost_center = center_map.get(trip.get("cost_center_id", ""), "N/A")
        
        ws.cell(row=row_idx, column=1, value=user_map.get(expense["user_id"], "N/A"))
        ws.cell(row=row_idx, column=2, value=cost_center)
        ws.cell(row=row_idx, column=3, value=expense["date"].strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=4, value=expense["amount"])
        ws.cell(row=row_idx, column=5, value=category_map.get(expense["category_id"], "N/A"))
        ws.cell(row=row_idx, column=6, value=expense["establishment"])
        ws.cell(row=row_idx, column=7, value=trip.get("name", "N/A"))
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"gastos_viaje_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Initialize default data
@app.on_event("startup")
async def startup_event():
    # Create default expense categories if none exist
    count = await db.expense_categories.count_documents({})
    if count == 0:
        default_categories = [
            "Transporte",
            "Alojamiento",
            "Comidas",
            "Combustible",
            "Parking",
            "Peajes",
            "Taxi/Uber",
            "Material de oficina",
            "Teléfono/Comunicaciones",
            "Otros gastos"
        ]
        
        for idx, name in enumerate(default_categories, 1):
            await db.expense_categories.insert_one({
                "category_id": f"category_{uuid.uuid4().hex[:12]}",
                "name": name,
                "active": True,
                "order": idx,
                "created_at": datetime.now(timezone.utc)
            })
        
        print(f"Created {len(default_categories)} default expense categories")

@app.get("/api/health")
async def health():
    return {"status": "ok"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
